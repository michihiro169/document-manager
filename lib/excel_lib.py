from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from src.excel.excel_specification import ExcelSpecification
import os

class ExcelLib():
    @classmethod
    def save(self, excel):
        wb = Workbook()
        for sheet in excel.getSheets():
            ws = wb.create_sheet(title = sheet.getName())

            # シートの列幅の設定
            for i, width in enumerate(sheet.getWidths()):
                ws.column_dimensions[ExcelSpecification.getAlphabet(i)].width = width

            # シートのオートフィルの設定
            if sheet.hasAutoFilter():
                ws.auto_filter.ref = sheet.getAutoFilter().getValue()

            # シートの行の処理
            for rowIndex, row in enumerate(sheet.getRows()):
                # シート行に値を追加
                values = []
                for cell in row:
                    values.append(cell.getValue())
                ws.append(values)

                # シート行の処理
                for cellIndex, cell in enumerate(row):
                    # データバリデーション
                    if cell.hasValidationList():
                        formula1 = ", ".join(cell.getDataValidation())
                        dv = DataValidation(
                            type="list",
                            formula1=f'"{formula1}"',
                            allow_blank=True,
                            showErrorMessage=True
                        )
                        alphabet = ExcelSpecification.getAlphabet(cellIndex)
                        dv.add(f"{alphabet}{rowIndex + 1}:{alphabet}{rowIndex + 1}")
                        ws.add_data_validation(dv)

                    # 書式
                    style = cell.getStyle()
                    alignment = style.getAlignment()
                    border = style.getBorder()
                    # 塗りつぶし
                    if style.hasFill():
                        ws[rowIndex + 1][cellIndex].fill = PatternFill(
                            fill_type = style.getFill().getType(),
                            fgColor   = style.getFill().getColor()
                        )
                    # 罫線
                    if border != None:
                        side = Side(style='thin', color='000000')
                        ws[rowIndex + 1][cellIndex].border = Border(
                            top = side if border.hasTop() else None,
                            bottom = side if border.hasBottom() else None,
                            left = side if border.hasLeft() else None,
                            right = side if border.hasRight() else None
                        )
                    # 配置
                    if alignment != None:
                        ws[rowIndex + 1][cellIndex].alignment = Alignment(
                            vertical = alignment.getVertical(),
                            wrapText = alignment.getWrapText()
                        )
                    # フォント
                    if style.hasFont():
                        ws[rowIndex + 1][cellIndex].font = Font(color=style.getFont().getColor())

            # シートの画像の処理
            for imageIndex, image in enumerate(sheet.getImages()):
                img = Image(image.getPath())
                img.width = image.getWidth()
                img.height = image.getHeight()
                ws.add_image(img, f'{ExcelSpecification.getAlphabet(image.getRowIndex())}{image.getColumnIndex()}')

        # Sheet1を削除
        sheet1 = wb.active
        wb.remove(sheet1)
        
        if not os.path.exists("./build"):
            os.makedirs("./build")

        wb.save(f"./build/{excel.getName()}")
